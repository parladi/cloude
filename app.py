"""
UBL-TR Fatura & Irsaliye Ayiklayici - Flask Web Uygulamasi
GIB e-Fatura ve e-Irsaliye dosyalarindan bilgileri ayiklar.
"""

import os
import io
import uuid
from threading import Lock
from flask import Flask, render_template, request, send_file, redirect, url_for, flash, session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ubl_parser import extract_from_zip_bytes, get_flat_invoice_data, get_all_line_items
from irsaliye_parser import (
    extract_from_zip_bytes as irsaliye_extract_from_zip_bytes,
    extract_from_xml_bytes_list,
    get_flat_irsaliye_data,
    get_all_despatch_lines,
)

# Gecici bellek cache - Excel icin yeniden upload gerekmez
_irsaliye_cache = {}
_cache_lock = Lock()

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB limit


# ===================== STIL YARDIMCILARI =====================

def get_excel_styles():
    """Excel icin ortak stil tanimlari."""
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return header_font, header_fill, header_alignment, thin_border


def write_excel_sheet(ws, data, styles):
    """Veri listesini Excel sayfasina yazar."""
    header_font, header_fill, header_alignment, thin_border = styles

    if not data:
        return

    all_keys = []
    for row in data:
        for key in row.keys():
            if key not in all_keys:
                all_keys.append(key)

    for col, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=col, value=key)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, key in enumerate(all_keys, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ''))
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    for col in range(1, len(all_keys) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'


# ===================== ANA SAYFA =====================

@app.route('/', methods=['GET'])
def index():
    """Ana sayfa - modul secimi."""
    return render_template('index.html')


# ===================== FATURA (INVOICE) =====================

@app.route('/upload', methods=['POST'])
def upload():
    """ZIP dosyasini yukle ve faturalari ayikla."""
    if 'zipfile' not in request.files:
        flash('Lutfen bir ZIP dosyasi secin.', 'error')
        return redirect(url_for('index'))

    file = request.files['zipfile']
    if file.filename == '':
        flash('Dosya secilmedi.', 'error')
        return redirect(url_for('index'))

    if not file.filename.lower().endswith('.zip'):
        flash('Lutfen .zip uzantili bir dosya yukleyin.', 'error')
        return redirect(url_for('index'))

    zip_bytes = file.read()
    invoices, errors = extract_from_zip_bytes(zip_bytes, file.filename)

    if not invoices and errors:
        flash(f'Hic fatura bulunamadi. {len(errors)} hata olustu.', 'error')
        return render_template('results.html', invoices=[], errors=errors,
                               flat_data=[], line_items=[], filename=file.filename)

    flat_data = get_flat_invoice_data(invoices)
    line_items = get_all_line_items(invoices)

    return render_template('results.html',
                           invoices=invoices,
                           flat_data=flat_data,
                           line_items=line_items,
                           errors=errors,
                           filename=file.filename)


@app.route('/export', methods=['POST'])
def export_excel():
    """Faturalari Excel dosyasina aktarir."""
    if 'zipfile' not in request.files:
        flash('Excel olusturmak icin ZIP dosyasi yukleyin.', 'error')
        return redirect(url_for('index'))

    file = request.files['zipfile']
    if file.filename == '' or not file.filename.lower().endswith('.zip'):
        flash('Gecerli bir ZIP dosyasi yukleyin.', 'error')
        return redirect(url_for('index'))

    zip_bytes = file.read()
    invoices, errors = extract_from_zip_bytes(zip_bytes, file.filename)

    if not invoices:
        flash('Ice aktarilacak fatura bulunamadi.', 'error')
        return redirect(url_for('index'))

    wb = create_excel_workbook(invoices)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    export_name = os.path.splitext(file.filename)[0] + '_faturalar.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=export_name
    )


def create_excel_workbook(invoices):
    """Fatura verilerinden Excel calisma kitabi olusturur."""
    wb = Workbook()
    styles = get_excel_styles()

    ws1 = wb.active
    ws1.title = 'Fatura Ozeti'
    flat_data = get_flat_invoice_data(invoices)
    write_excel_sheet(ws1, flat_data, styles)

    ws2 = wb.create_sheet('Fatura Kalemleri')
    line_items = get_all_line_items(invoices)
    write_excel_sheet(ws2, line_items, styles)

    return wb


# ===================== IRSALIYE (DESPATCH ADVICE) =====================

@app.route('/irsaliye', methods=['GET'])
def irsaliye_index():
    """Irsaliye yukle sayfasi."""
    return render_template('irsaliye_index.html')


@app.route('/irsaliye/upload', methods=['POST'])
def irsaliye_upload():
    """ZIP veya XML dosyalarini yukle ve irsaliyeleri ayikla."""
    uploaded_files = request.files.getlist('xmlfiles')

    if not uploaded_files or all(f.filename == '' for f in uploaded_files):
        flash('Lutfen dosya secin.', 'error')
        return redirect(url_for('irsaliye_index'))

    irsaliyeler = []
    errors = []
    first_filename = ''

    for file in uploaded_files:
        if file.filename == '':
            continue

        if not first_filename:
            first_filename = file.filename

        file_bytes = file.read()

        if file.filename.lower().endswith('.zip'):
            # ZIP dosyasi - icindeki XML'leri ayikla
            irs_list, err_list = irsaliye_extract_from_zip_bytes(file_bytes, file.filename)
            irsaliyeler.extend(irs_list)
            errors.extend(err_list)
        elif file.filename.lower().endswith('.xml'):
            # Tek XML dosyasi
            irs_list, err_list = extract_from_xml_bytes_list([(file.filename, file_bytes)])
            irsaliyeler.extend(irs_list)
            errors.extend(err_list)
        else:
            errors.append({'dosya': file.filename, 'hata': 'Desteklenmeyen dosya formati. XML veya ZIP yukleyin.'})

    if not irsaliyeler and errors:
        flash(f'Hic irsaliye bulunamadi. {len(errors)} hata olustu.', 'error')

    flat_data = get_flat_irsaliye_data(irsaliyeler)
    line_items = get_all_despatch_lines(irsaliyeler)

    # Sütun başlıklarını Python'da hesapla (template'e geçir)
    flat_keys = []
    for row in flat_data:
        for key in row.keys():
            if key not in flat_keys:
                flat_keys.append(key)

    line_keys = []
    for row in line_items:
        for key in row.keys():
            if key not in line_keys:
                line_keys.append(key)

    # Excel icin veriyi cache'le - re-upload gerekmez
    cache_key = str(uuid.uuid4())
    with _cache_lock:
        _irsaliye_cache[cache_key] = irsaliyeler
        if len(_irsaliye_cache) > 50:
            oldest = list(_irsaliye_cache.keys())[0]
            del _irsaliye_cache[oldest]

    return render_template('irsaliye_results.html',
                           irsaliyeler=irsaliyeler,
                           flat_data=flat_data,
                           line_items=line_items,
                           flat_keys=flat_keys,
                           line_keys=line_keys,
                           errors=errors,
                           filename=first_filename,
                           file_count=len(uploaded_files),
                           cache_key=cache_key)


@app.route('/irsaliye/download/<cache_key>')
def irsaliye_download_excel(cache_key):
    """Cache'teki irsaliye verisini Excel olarak indirir (re-upload gerekmez)."""
    with _cache_lock:
        irsaliyeler = _irsaliye_cache.get(cache_key)

    if not irsaliyeler:
        flash('Oturum suresi doldu. Lutfen dosyalari tekrar yukleyin.', 'error')
        return redirect(url_for('irsaliye_index'))

    wb = create_irsaliye_excel(irsaliyeler)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='irsaliyeler.xlsx'
    )


@app.route('/irsaliye/export', methods=['POST'])
def irsaliye_export_excel():
    """Irsaliyeleri Excel dosyasina aktarir (dosya yeniden yuklenerek)."""
    uploaded_files = request.files.getlist('xmlfiles')

    if not uploaded_files or all(f.filename == '' for f in uploaded_files):
        flash('Excel olusturmak icin dosya yukleyin.', 'error')
        return redirect(url_for('irsaliye_index'))

    irsaliyeler = []
    errors = []
    first_filename = 'irsaliyeler'

    for file in uploaded_files:
        if file.filename == '':
            continue
        if first_filename == 'irsaliyeler':
            first_filename = os.path.splitext(file.filename)[0]

        file_bytes = file.read()

        if file.filename.lower().endswith('.zip'):
            irs_list, err_list = irsaliye_extract_from_zip_bytes(file_bytes, file.filename)
            irsaliyeler.extend(irs_list)
            errors.extend(err_list)
        elif file.filename.lower().endswith('.xml'):
            irs_list, err_list = extract_from_xml_bytes_list([(file.filename, file_bytes)])
            irsaliyeler.extend(irs_list)
            errors.extend(err_list)

    if not irsaliyeler:
        flash('Ice aktarilacak irsaliye bulunamadi.', 'error')
        return redirect(url_for('irsaliye_index'))

    wb = create_irsaliye_excel(irsaliyeler)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    export_name = first_filename + '_irsaliyeler.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=export_name
    )


def create_irsaliye_excel(irsaliyeler):
    """Irsaliye verilerinden Excel calisma kitabi olusturur."""
    wb = Workbook()
    styles = get_excel_styles()

    ws1 = wb.active
    ws1.title = 'Irsaliye Ozeti'
    flat_data = get_flat_irsaliye_data(irsaliyeler)
    write_excel_sheet(ws1, flat_data, styles)

    ws2 = wb.create_sheet('Irsaliye Kalemleri')
    line_items = get_all_despatch_lines(irsaliyeler)
    write_excel_sheet(ws2, line_items, styles)

    return wb


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
